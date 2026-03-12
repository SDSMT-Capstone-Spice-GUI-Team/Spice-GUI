"""Tests for Bill of Materials (BOM) export functionality."""

import csv
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook
from simulation.bom_exporter import export_bom_csv, export_bom_excel, generate_bom, write_bom_csv
from tests.conftest import make_component


def _build_components(*specs):
    """Build a components dict from (type, id, value) tuples."""
    return {s[1]: make_component(s[0], s[1], s[2]) for s in specs}


class TestGenerateBom:
    def test_basic_bom(self):
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Capacitor", "C1", "10u"),
        )
        bom = generate_bom(components)
        assert bom["total_count"] == 2
        assert len(bom["items"]) == 2
        assert len(bom["summary"]) == 2

    def test_excludes_ground(self):
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Ground", "GND1", "0V"),
        )
        bom = generate_bom(components)
        assert bom["total_count"] == 1
        ids = [item["id"] for item in bom["items"]]
        assert "GND1" not in ids

    def test_sorted_by_type_then_id(self):
        components = _build_components(
            ("Resistor", "R2", "2k"),
            ("Capacitor", "C1", "10u"),
            ("Resistor", "R1", "1k"),
        )
        bom = generate_bom(components)
        ids = [item["id"] for item in bom["items"]]
        assert ids == ["C1", "R1", "R2"]

    def test_quantity_summary(self):
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Resistor", "R2", "1k"),
            ("Resistor", "R3", "2k"),
            ("Capacitor", "C1", "10u"),
        )
        bom = generate_bom(components)
        summary = {(r["type"], r["value"]): r["quantity"] for r in bom["summary"]}
        assert summary[("Resistor", "1k")] == 2
        assert summary[("Resistor", "2k")] == 1
        assert summary[("Capacitor", "10u")] == 1

    def test_empty_circuit(self):
        bom = generate_bom({})
        assert bom["total_count"] == 0
        assert bom["items"] == []
        assert bom["summary"] == []

    def test_circuit_name(self):
        bom = generate_bom({}, circuit_name="test_circuit.json")
        assert bom["circuit_name"] == "test_circuit.json"

    def test_ground_only_circuit(self):
        components = _build_components(("Ground", "GND1", "0V"))
        bom = generate_bom(components)
        assert bom["total_count"] == 0


class TestExportBomCsv:
    def test_produces_valid_csv(self):
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Capacitor", "C1", "10u"),
        )
        content = export_bom_csv(components)
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        assert len(rows) > 0

    def test_contains_header_row(self):
        components = _build_components(("Resistor", "R1", "1k"))
        content = export_bom_csv(components)
        assert "Component ID" in content
        assert "Type" in content
        assert "Value" in content

    def test_contains_component_data(self):
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Capacitor", "C1", "10u"),
        )
        content = export_bom_csv(components)
        assert "R1" in content
        assert "Resistor" in content
        assert "1k" in content
        assert "C1" in content

    def test_contains_quantity_summary(self):
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Resistor", "R2", "1k"),
        )
        content = export_bom_csv(components)
        assert "Quantity Summary" in content
        assert "Total Components" in content

    def test_empty_circuit_produces_headers(self):
        content = export_bom_csv({})
        assert "Component ID" in content
        assert "Bill of Materials" in content

    def test_circuit_name_in_metadata(self):
        content = export_bom_csv({}, circuit_name="my_circuit")
        assert "my_circuit" in content


class TestExportBomExcel:
    def test_creates_valid_xlsx(self, tmp_path):
        path = tmp_path / "bom.xlsx"
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Capacitor", "C1", "10u"),
        )
        export_bom_excel(components, str(path))
        wb = load_workbook(str(path))
        assert "Summary" in wb.sheetnames
        assert "Component List" in wb.sheetnames
        assert "Quantity Summary" in wb.sheetnames

    def test_component_list_content(self, tmp_path):
        path = tmp_path / "bom.xlsx"
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Resistor", "R2", "2k"),
        )
        export_bom_excel(components, str(path))
        wb = load_workbook(str(path))
        ws = wb["Component List"]
        # Header + 2 data rows
        assert ws.max_row == 3
        assert ws.cell(row=1, column=1).value == "Component ID"
        assert ws.cell(row=2, column=1).value in ("R1", "R2")

    def test_quantity_summary_content(self, tmp_path):
        path = tmp_path / "bom.xlsx"
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Resistor", "R2", "1k"),
            ("Capacitor", "C1", "10u"),
        )
        export_bom_excel(components, str(path))
        wb = load_workbook(str(path))
        ws = wb["Quantity Summary"]
        # Header + 2 summary rows (Capacitor 10u, Resistor 1k)
        assert ws.max_row == 3

    def test_summary_has_total(self, tmp_path):
        path = tmp_path / "bom.xlsx"
        components = _build_components(("Resistor", "R1", "1k"))
        export_bom_excel(components, str(path), circuit_name="test.json")
        wb = load_workbook(str(path))
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Total Components" in values

    def test_empty_circuit(self, tmp_path):
        path = tmp_path / "bom.xlsx"
        export_bom_excel({}, str(path))
        wb = load_workbook(str(path))
        assert "Component List" in wb.sheetnames
        ws = wb["Component List"]
        # Header only
        assert ws.max_row == 1

    def test_excludes_ground(self, tmp_path):
        path = tmp_path / "bom.xlsx"
        components = _build_components(
            ("Resistor", "R1", "1k"),
            ("Ground", "GND1", "0V"),
        )
        export_bom_excel(components, str(path))
        wb = load_workbook(str(path))
        ws = wb["Component List"]
        ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "GND1" not in ids


class TestWriteBomCsv:
    def test_writes_file(self, tmp_path):
        path = tmp_path / "bom.csv"
        write_bom_csv("a,b,c\n1,2,3\n", str(path))
        assert path.exists()
        assert path.read_text() == "a,b,c\n1,2,3\n"


class TestNoQtDependencies:
    def test_no_pyqt_imports(self):
        import simulation.bom_exporter as mod

        source = Path(mod.__file__).read_text(encoding="utf-8")
        assert "PyQt" not in source
        assert "QtCore" not in source
        assert "QtWidgets" not in source
