"""
simulation/bom_exporter.py

Export Bill of Materials (BOM) listing all components with ID, type, value,
and quantity summaries grouped by type+value.
No Qt dependencies — file dialog is the view's responsibility.
"""

import csv
import io
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generate_bom(components, circuit_name=""):
    """Generate a BOM data structure from circuit components.

    Args:
        components: dict mapping component_id -> ComponentData
        circuit_name: optional circuit filename for metadata

    Returns:
        dict with keys:
            'items': list of dicts with 'id', 'type', 'value' (sorted by type, then id)
            'summary': list of dicts with 'type', 'value', 'quantity' (sorted by type, then value)
            'circuit_name': str
            'date': str
            'total_count': int (total components excluding Ground)
    """
    # Filter out Ground components — they aren't physical parts
    items = []
    for comp_id, comp in sorted(components.items()):
        if comp.component_type == "Ground":
            continue
        items.append(
            {
                "id": comp.component_id,
                "type": comp.component_type,
                "value": comp.value,
            }
        )

    # Sort by type, then by component ID
    items.sort(key=lambda x: (x["type"], x["id"]))

    # Build quantity summary grouped by (type, value)
    counter = Counter((item["type"], item["value"]) for item in items)
    summary = []
    for (comp_type, value), qty in sorted(counter.items()):
        summary.append({"type": comp_type, "value": value, "quantity": qty})

    return {
        "items": items,
        "summary": summary,
        "circuit_name": circuit_name,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_count": len(items),
    }


def export_bom_csv(components, circuit_name=""):
    """Export BOM as a CSV string.

    Args:
        components: dict mapping component_id -> ComponentData
        circuit_name: optional circuit filename

    Returns:
        str: CSV content
    """
    bom = generate_bom(components, circuit_name)
    output = io.StringIO()
    writer = csv.writer(output)

    # Metadata header
    writer.writerow(["# Bill of Materials"])
    writer.writerow(["# Date", bom["date"]])
    if bom["circuit_name"]:
        writer.writerow(["# Circuit", bom["circuit_name"]])
    writer.writerow([])

    # Component list
    writer.writerow(["Component ID", "Type", "Value"])
    for item in bom["items"]:
        writer.writerow([item["id"], item["type"], item["value"]])
    writer.writerow([])

    # Quantity summary
    writer.writerow(["# Quantity Summary"])
    writer.writerow(["Type", "Value", "Quantity"])
    for row in bom["summary"]:
        writer.writerow([row["type"], row["value"], row["quantity"]])
    writer.writerow([])
    writer.writerow(["# Total Components", bom["total_count"]])

    return output.getvalue()


def export_bom_excel(components, filepath, circuit_name=""):
    """Export BOM to an Excel workbook.

    Args:
        components: dict mapping component_id -> ComponentData
        filepath: path to write the .xlsx file
        circuit_name: optional circuit filename for metadata
    """
    bom = generate_bom(components, circuit_name)
    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def style_header(ws, row_num=1):
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Bill of Materials"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(["Date", bom["date"]])
    if bom["circuit_name"]:
        ws_summary.append(["Circuit", bom["circuit_name"]])
    ws_summary.append(["Total Components", bom["total_count"]])
    for row in ws_summary.iter_rows(min_row=3, max_col=1):
        row[0].font = Font(bold=True)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30

    # --- Component List sheet ---
    ws_items = wb.create_sheet("Component List")
    ws_items.append(["Component ID", "Type", "Value"])
    style_header(ws_items)
    for item in bom["items"]:
        ws_items.append([item["id"], item["type"], item["value"]])
    ws_items.column_dimensions["A"].width = 18
    ws_items.column_dimensions["B"].width = 20
    ws_items.column_dimensions["C"].width = 15

    # --- Quantity Summary sheet ---
    ws_qty = wb.create_sheet("Quantity Summary")
    ws_qty.append(["Type", "Value", "Quantity"])
    style_header(ws_qty)
    for row in bom["summary"]:
        ws_qty.append([row["type"], row["value"], row["quantity"]])
    ws_qty.column_dimensions["A"].width = 20
    ws_qty.column_dimensions["B"].width = 15
    ws_qty.column_dimensions["C"].width = 12

    wb.save(filepath)


def write_bom_csv(csv_content, filepath):
    """Write BOM CSV content string to a file.

    Args:
        csv_content: str from export_bom_csv()
        filepath: path to write to
    """
    with open(filepath, "w", newline="") as f:
        f.write(csv_content)
