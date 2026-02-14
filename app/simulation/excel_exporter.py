"""
simulation/excel_exporter.py

Export simulation results to Excel (.xlsx) format.
No Qt dependencies — file dialog is the view's responsibility.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _add_metadata_sheet(wb, analysis_type, circuit_name=""):
    """Add a Summary sheet with circuit metadata."""
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True)
    ws.append(["Circuit Report Summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Analysis Type", analysis_type])
    ws.append(["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    if circuit_name:
        ws.append(["Circuit", circuit_name])
    for row in ws.iter_rows(min_row=3, max_col=1):
        row[0].font = header_font
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    return ws


def _style_header_row(ws, row_num=1):
    """Apply header styling to the first row of a worksheet."""
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def export_to_excel(results, analysis_type, filepath, circuit_name=""):
    """Export simulation results to an Excel workbook.

    Args:
        results: simulation result data (format depends on analysis_type)
        analysis_type: one of "DC Operating Point", "DC Sweep", "AC Sweep",
                       "Transient", "Noise"
        filepath: path to write the .xlsx file
        circuit_name: optional circuit filename for metadata
    """
    wb = Workbook()
    _add_metadata_sheet(wb, analysis_type, circuit_name)

    if analysis_type == "DC Operating Point":
        _export_op(wb, results)
    elif analysis_type == "DC Sweep":
        _export_dc_sweep(wb, results)
    elif analysis_type == "AC Sweep":
        _export_ac(wb, results)
    elif analysis_type == "Transient":
        _export_transient(wb, results)
    elif analysis_type == "Noise":
        _export_noise(wb, results)

    wb.save(filepath)


def _export_op(wb, node_voltages):
    """Export DC Operating Point results."""
    ws = wb.create_sheet("DC Operating Point")
    ws.append(["Node", "Voltage (V)"])
    _style_header_row(ws)
    for node, voltage in sorted(node_voltages.items()):
        ws.append([node, voltage])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def _export_dc_sweep(wb, sweep_data):
    """Export DC Sweep results."""
    ws = wb.create_sheet("DC Sweep")
    headers = sweep_data.get("headers", [])
    ws.append(headers)
    _style_header_row(ws)
    for row in sweep_data.get("data", []):
        ws.append(row)
    for i in range(len(headers)):
        col_letter = chr(ord("A") + i) if i < 26 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = 15


def _export_ac(wb, ac_data):
    """Export AC Sweep results with magnitude and phase columns."""
    ws = wb.create_sheet("AC Sweep")
    frequencies = ac_data.get("frequencies", [])
    magnitude = ac_data.get("magnitude", {})
    phase = ac_data.get("phase", {})

    all_nodes = sorted(set(magnitude.keys()) | set(phase.keys()))
    headers = ["Frequency (Hz)"]
    for node in all_nodes:
        if node in magnitude:
            headers.append(f"|V({node})| (dB)")
        if node in phase:
            headers.append(f"phase({node}) (deg)")
    ws.append(headers)
    _style_header_row(ws)

    for i, freq in enumerate(frequencies):
        row = [freq]
        for node in all_nodes:
            if node in magnitude:
                row.append(magnitude[node][i] if i < len(magnitude[node]) else "")
            if node in phase:
                row.append(phase[node][i] if i < len(phase[node]) else "")
        ws.append(row)

    for j in range(len(headers)):
        col_letter = chr(ord("A") + j) if j < 26 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = 18


def _export_transient(wb, tran_data):
    """Export Transient analysis results."""
    ws = wb.create_sheet("Transient")
    if not tran_data:
        ws.append(["No data"])
        return

    headers = list(tran_data[0].keys())
    # Add units to known headers
    display_headers = []
    for h in headers:
        if h == "time":
            display_headers.append("Time (s)")
        else:
            display_headers.append(f"{h} (V)")
    ws.append(display_headers)
    _style_header_row(ws)

    for row in tran_data:
        ws.append([row[h] for h in headers])

    for j in range(len(headers)):
        col_letter = chr(ord("A") + j) if j < 26 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = 15


def _export_noise(wb, noise_data):
    """Export Noise analysis results."""
    ws = wb.create_sheet("Noise")
    frequencies = noise_data.get("frequencies", [])
    onoise = noise_data.get("onoise_spectrum", [])
    inoise = noise_data.get("inoise_spectrum", [])

    headers = ["Frequency (Hz)"]
    if onoise:
        headers.append("Output Noise (V/sqrt(Hz))")
    if inoise:
        headers.append("Input Noise (V/sqrt(Hz))")
    ws.append(headers)
    _style_header_row(ws)

    for i, freq in enumerate(frequencies):
        row = [freq]
        if onoise:
            row.append(onoise[i] if i < len(onoise) else "")
        if inoise:
            row.append(inoise[i] if i < len(inoise) else "")
        ws.append(row)

    for j in range(len(headers)):
        col_letter = chr(ord("A") + j) if j < 26 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = 25
