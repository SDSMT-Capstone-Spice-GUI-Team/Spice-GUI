"""Tests for Excel (.xlsx) export functionality."""

import pytest
from openpyxl import load_workbook
from simulation.excel_exporter import export_to_excel


class TestExportOpToExcel:
    def test_creates_valid_xlsx(self, tmp_path):
        path = tmp_path / "op.xlsx"
        voltages = {"1": 5.0, "2": 3.3}
        export_to_excel(voltages, "DC Operating Point", str(path))
        wb = load_workbook(str(path))
        assert "Summary" in wb.sheetnames
        assert "DC Operating Point" in wb.sheetnames

    def test_contains_all_nodes(self, tmp_path):
        path = tmp_path / "op.xlsx"
        voltages = {"out": 2.5, "in": 5.0}
        export_to_excel(voltages, "DC Operating Point", str(path))
        wb = load_workbook(str(path))
        ws = wb["DC Operating Point"]
        nodes = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "in" in nodes
        assert "out" in nodes

    def test_header_row(self, tmp_path):
        path = tmp_path / "op.xlsx"
        export_to_excel({"1": 1.0}, "DC Operating Point", str(path))
        wb = load_workbook(str(path))
        ws = wb["DC Operating Point"]
        assert ws.cell(row=1, column=1).value == "Node"
        assert ws.cell(row=1, column=2).value == "Voltage (V)"

    def test_sorted_by_node(self, tmp_path):
        path = tmp_path / "op.xlsx"
        export_to_excel({"z": 1.0, "a": 2.0}, "DC Operating Point", str(path))
        wb = load_workbook(str(path))
        ws = wb["DC Operating Point"]
        assert ws.cell(row=2, column=1).value == "a"
        assert ws.cell(row=3, column=1).value == "z"


class TestExportDcSweepToExcel:
    def test_creates_sheet(self, tmp_path):
        path = tmp_path / "dc.xlsx"
        data = {
            "headers": ["Index", "v-sweep", "v(1)"],
            "data": [[0, 0.0, 0.0], [1, 1.0, 0.5]],
        }
        export_to_excel(data, "DC Sweep", str(path))
        wb = load_workbook(str(path))
        assert "DC Sweep" in wb.sheetnames

    def test_data_rows(self, tmp_path):
        path = tmp_path / "dc.xlsx"
        data = {
            "headers": ["Index", "v-sweep", "v(1)"],
            "data": [[0, 0.0, 0.0], [1, 1.0, 0.5], [2, 2.0, 1.0]],
        }
        export_to_excel(data, "DC Sweep", str(path))
        wb = load_workbook(str(path))
        ws = wb["DC Sweep"]
        # header + 3 data rows
        assert ws.max_row == 4

    def test_empty_data(self, tmp_path):
        path = tmp_path / "dc.xlsx"
        data = {"headers": ["Index"], "data": []}
        export_to_excel(data, "DC Sweep", str(path))
        wb = load_workbook(str(path))
        assert "DC Sweep" in wb.sheetnames


class TestExportAcToExcel:
    def test_creates_sheet(self, tmp_path):
        path = tmp_path / "ac.xlsx"
        data = {
            "frequencies": [100.0, 1000.0],
            "magnitude": {"out": [0.5, 0.3]},
            "phase": {"out": [-45.0, -90.0]},
        }
        export_to_excel(data, "AC Sweep", str(path))
        wb = load_workbook(str(path))
        assert "AC Sweep" in wb.sheetnames

    def test_mag_and_phase_columns(self, tmp_path):
        path = tmp_path / "ac.xlsx"
        data = {
            "frequencies": [100.0],
            "magnitude": {"out": [0.5]},
            "phase": {"out": [-45.0]},
        }
        export_to_excel(data, "AC Sweep", str(path))
        wb = load_workbook(str(path))
        ws = wb["AC Sweep"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Frequency (Hz)" in headers
        assert any("out" in h and "dB" in h for h in headers if h)
        assert any("out" in h and "deg" in h for h in headers if h)

    def test_data_count(self, tmp_path):
        path = tmp_path / "ac.xlsx"
        data = {
            "frequencies": [100.0, 1000.0, 10000.0],
            "magnitude": {"out": [0.5, 0.3, 0.1]},
            "phase": {"out": [-45.0, -90.0, -135.0]},
        }
        export_to_excel(data, "AC Sweep", str(path))
        wb = load_workbook(str(path))
        ws = wb["AC Sweep"]
        assert ws.max_row == 4  # header + 3 data

    def test_empty_data(self, tmp_path):
        path = tmp_path / "ac.xlsx"
        data = {"frequencies": [], "magnitude": {}, "phase": {}}
        export_to_excel(data, "AC Sweep", str(path))
        wb = load_workbook(str(path))
        assert "AC Sweep" in wb.sheetnames


class TestExportTransientToExcel:
    def test_creates_sheet(self, tmp_path):
        path = tmp_path / "tran.xlsx"
        data = [{"time": 0.0, "out": 0.0}, {"time": 0.001, "out": 2.5}]
        export_to_excel(data, "Transient", str(path))
        wb = load_workbook(str(path))
        assert "Transient" in wb.sheetnames

    def test_headers_have_units(self, tmp_path):
        path = tmp_path / "tran.xlsx"
        data = [{"time": 0.0, "v1": 1.0}]
        export_to_excel(data, "Transient", str(path))
        wb = load_workbook(str(path))
        ws = wb["Transient"]
        assert ws.cell(row=1, column=1).value == "Time (s)"
        assert ws.cell(row=1, column=2).value == "v1 (V)"

    def test_data_rows(self, tmp_path):
        path = tmp_path / "tran.xlsx"
        data = [{"time": i * 0.001, "v1": float(i)} for i in range(5)]
        export_to_excel(data, "Transient", str(path))
        wb = load_workbook(str(path))
        ws = wb["Transient"]
        assert ws.max_row == 6  # header + 5

    def test_empty_data(self, tmp_path):
        path = tmp_path / "tran.xlsx"
        export_to_excel([], "Transient", str(path))
        wb = load_workbook(str(path))
        assert "Transient" in wb.sheetnames


class TestExportNoiseToExcel:
    def test_creates_sheet(self, tmp_path):
        path = tmp_path / "noise.xlsx"
        data = {
            "frequencies": [100.0, 1000.0],
            "onoise_spectrum": [1e-6, 2e-6],
            "inoise_spectrum": [3e-6, 4e-6],
        }
        export_to_excel(data, "Noise", str(path))
        wb = load_workbook(str(path))
        assert "Noise" in wb.sheetnames

    def test_headers(self, tmp_path):
        path = tmp_path / "noise.xlsx"
        data = {
            "frequencies": [100.0],
            "onoise_spectrum": [1e-6],
            "inoise_spectrum": [3e-6],
        }
        export_to_excel(data, "Noise", str(path))
        wb = load_workbook(str(path))
        ws = wb["Noise"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Frequency (Hz)" in headers
        assert "Output Noise (V/sqrt(Hz))" in headers
        assert "Input Noise (V/sqrt(Hz))" in headers

    def test_data_count(self, tmp_path):
        path = tmp_path / "noise.xlsx"
        data = {
            "frequencies": [100.0, 1000.0, 10000.0],
            "onoise_spectrum": [1e-6, 2e-6, 3e-6],
            "inoise_spectrum": [4e-6, 5e-6, 6e-6],
        }
        export_to_excel(data, "Noise", str(path))
        wb = load_workbook(str(path))
        ws = wb["Noise"]
        assert ws.max_row == 4  # header + 3


class TestMetadataSheet:
    def test_summary_sheet_exists(self, tmp_path):
        path = tmp_path / "meta.xlsx"
        export_to_excel({"1": 1.0}, "DC Operating Point", str(path))
        wb = load_workbook(str(path))
        assert "Summary" in wb.sheetnames

    def test_circuit_name_in_summary(self, tmp_path):
        path = tmp_path / "meta.xlsx"
        export_to_excel(
            {"1": 1.0}, "DC Operating Point", str(path), circuit_name="test.json"
        )
        wb = load_workbook(str(path))
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "test.json" in values

    def test_analysis_type_in_summary(self, tmp_path):
        path = tmp_path / "meta.xlsx"
        export_to_excel({"1": 1.0}, "DC Operating Point", str(path))
        wb = load_workbook(str(path))
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "DC Operating Point" in values


class TestNoQtDependencies:
    def test_no_pyqt_imports(self):
        import simulation.excel_exporter as mod

        source = open(mod.__file__).read()
        assert "PyQt" not in source
        assert "QtCore" not in source
        assert "QtWidgets" not in source
